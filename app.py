# -*- coding: utf-8 -*-
import base64
import requests
import os
import io, csv
from datetime import datetime, date, timedelta
from functools import wraps
from sqlalchemy import func, case, text, cast, Date, or_
from flask import Flask, render_template, request, redirect, url_for, session, flash, abort, make_response
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
import sync_service
import threading
import time
from zoneinfo import ZoneInfo
from xhtml2pdf import pisa
from flask_wtf.csrf import CSRFProtect  # Protección CSRF
# ----------------------------
# CONFIGURACIÓN ZONA HORARIA
# ----------------------------
ARG_TZ = ZoneInfo("America/Argentina/Buenos_Aires")

def get_arg_now():
    """Devuelve datetime actual en Argentina"""
    return datetime.now(ARG_TZ)

def get_arg_today():
    """Devuelve date actual en Argentina"""
    return datetime.now(ARG_TZ).date()

# ----------------------------
# Configuración por .env
# ----------------------------
try:
    from dotenv import load_dotenv  # type: ignore
    load_dotenv()
except Exception:
    pass

# Variables de ambiente
SECRET_KEY   = os.getenv("FLASK_SECRET_KEY") or os.getenv("SECRET_KEY", "clave-super-secreta-para-dev")
TPL_EXT      = os.getenv("TPL_EXT", ".html")
PASSWORD_LOG = os.getenv("PASSWORD_LOG", os.path.join("Data", "passwords.log"))
ADMIN_USER   = os.getenv("ADMIN_USER", "admin")
ADMIN_PASS   = os.getenv("ADMIN_PASS", "admin123")
DB_SCHEMA    = os.getenv("DB_SCHEMA", "transportistas")

# Solo PostgreSQL (sin fallback)
DATABASE_URL = (os.getenv("DATABASE_URL") or "").strip()
if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL no está definido. La app requiere PostgreSQL.")

if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://","postgresql+psycopg://", 1)
elif DATABASE_URL.startswith("postgresql://"):
    DATABASE_URL = "postgresql+psycopg://" + DATABASE_URL.split("://", 1)[1]

if DATABASE_URL.startswith("postgresql+psycopg://"):
    sep = '&' if '?' in DATABASE_URL else '?'
    DATABASE_URL = f"{DATABASE_URL}{sep}options=-csearch_path%3D{DB_SCHEMA},public"


# ----------------------------
# Inicialización Flask / DB
# ----------------------------
app = Flask(__name__, template_folder="templates")
app.secret_key = SECRET_KEY

# Inicializar protección CSRF globalmente
csrf = CSRFProtect(app) 

app.jinja_env.globals.update(get_arg_today=get_arg_today)

app.config.update(
    SQLALCHEMY_DATABASE_URI=DATABASE_URL,
    SQLALCHEMY_TRACK_MODIFICATIONS=False,
    SQLALCHEMY_ENGINE_OPTIONS={
        "pool_size": 5,
        "max_overflow": 5,
        "pool_pre_ping": True,
        "pool_recycle": 1800,
    },
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=(os.getenv("FLASK_ENV") == "production" or os.getenv("USE_HTTPS") == "True")
)

db = SQLAlchemy(app)

# ----------------------------
# Mails
# ----------------------------

app.config['MAIL_SERVER'] = os.getenv('MAIL_SERVER')
app.config['MAIL_PORT'] = int(os.getenv('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.getenv('MAIL_USE_TLS') == 'True'
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.getenv('MAIL_DEFAULT_SENDER')

_safe = DATABASE_URL.split("@", 1)[-1] if "@" in DATABASE_URL else DATABASE_URL
app.logger.info(
    "DB URI efectiva → postgresql://***:***@" + _safe
    if "postgresql+psycopg" in DATABASE_URL else
    "DB URI efectiva → " + _safe
)

os.makedirs(os.path.dirname(PASSWORD_LOG), exist_ok=True)

def tpl(name: str) -> str:
    return f"{name}{TPL_EXT}"

def _resolve_range(data: dict):
    rng = (data.get('range') or '').strip().lower()
    dfrom = data.get('from')
    dto   = data.get('to')
    if dfrom and dto:
        try:
            start = date.fromisoformat(dfrom)
            end   = date.fromisoformat(dto)
            if end < start:
                start, end = end, start
            return start, end
        except Exception:
            pass
    hoy = get_arg_today()
    if rng in ('', 'today'):
        return hoy, hoy
    if rng == 'week':
        return hoy - timedelta(days=6), hoy
    if rng == 'month':
        start = hoy.replace(day=1)
        if start.month == 12:
            end = start.replace(year=start.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end = start.replace(month=start.month + 1, day=1) - timedelta(days=1)
        return start, end
    return hoy, hoy

def get_maestro_id(user_id):
    u = db.session.get(User, user_id)
    if not u:
        return None
    if u.parent_id:
        return u.parent_id
    return u.id

def get_family_ids(user_id):
    maestro_id = get_maestro_id(user_id)
    if not maestro_id:
        return []
    sub_users = db.session.query(User.id).filter_by(parent_id=maestro_id).all()
    ids = [maestro_id] + [u.id for u in sub_users]
    return ids

def norm_username(s: str) -> str:
    return (s or "").strip().lower()

def calcular_cuil(dni_str: str, gender: str | None) -> str:
    digits = ''.join(ch for ch in (dni_str or '') if ch.isdigit()).zfill(8)
    if len(digits) != 8:
        return ""
    g = (gender or "").upper()
    pref = 20 if g == "M" else (27 if g == "F" else 23)
    weights = [5,4,3,2,7,6,5,4,3,2]
    def calc_dv(prefijo: int):
        base = f"{prefijo:02d}{digits}"
        total = sum(int(d)*w for d, w in zip(base, weights))
        dv = 11 - (total % 11)
        if dv == 11: return 0
        if dv == 10: return None
        return dv
    dv = calc_dv(pref)
    if dv is None:
        pref = 23
        dv = calc_dv(pref)
        if dv is None:
            pref = 24
            dv = calc_dv(pref)
            if dv is None:
                return ""
    return f"{pref:02d}{digits}{dv}"

def get_config():
    conf = SystemConfig.query.first()
    if not conf:
        conf = SystemConfig()
        db.session.add(conf)
        db.session.commit()
    return conf

# ----------------------------
# MODELOS
# ----------------------------
class User(db.Model):
    __tablename__ = "user"
    id             = db.Column(db.Integer, primary_key=True)
    username       = db.Column(db.String(50), unique=True, nullable=False, index=True)
    password_hash  = db.Column(db.String(512), nullable=False)
    tipo           = db.Column(db.String(20), nullable=False)
    email          = db.Column(db.String(120), nullable=True)
    custom_price   = db.Column(db.Float, default=0.0)
    cert_type      = db.Column(db.String(20), default='llegada')
    parent_id      = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    sub_users      = db.relationship("User", backref=db.backref("parent", remote_side=[id]),lazy="dynamic")
    payment_days   = db.Column(db.Integer, default=30)

    shipments_sent = db.relationship("Shipment", foreign_keys="Shipment.transportista_id", backref="transportista", cascade="all, delete-orphan", lazy="dynamic")
    shipments_received = db.relationship("Shipment", foreign_keys="Shipment.arenera_id", backref="arenera", cascade="all, delete-orphan", lazy="dynamic")
    quotas = db.relationship("Quota", foreign_keys="Quota.transportista_id", backref="transportista_user", cascade="all, delete-orphan", lazy="dynamic")
    shipments_operated = db.relationship("Shipment", foreign_keys="Shipment.operador_id", backref="operador", lazy="dynamic")

class Shipment(db.Model):
    __tablename__ = "shipment"
    id               = db.Column(db.Integer, primary_key=True)
    transportista_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    arenera_id       = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    operador_id      = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True, default=0)
    date             = db.Column(db.Date, nullable=False, index=True)
    
    chofer           = db.Column(db.String(100), nullable=False)
    dni              = db.Column(db.String(50), nullable=False)
    gender           = db.Column(db.String(10), nullable=False)
    tipo             = db.Column(db.String(20), nullable=False)
    tractor          = db.Column(db.String(20), nullable=False)
    trailer          = db.Column(db.String(20), nullable=False)
    
    status           = db.Column(db.String(20), nullable=False, default="En viaje", index=True)
    remito_arenera    = db.Column(db.String(50), nullable=True, index=True)
    peso_neto_arenera = db.Column(db.Float, nullable=True)

    sbe_remito        = db.Column(db.String(50), nullable=True)
    sbe_peso_neto     = db.Column(db.Float, nullable=True)
    sbe_fecha_salida  = db.Column(db.DateTime, nullable=True)
    sbe_fecha_llegada = db.Column(db.DateTime, nullable=True)
    sbe_patente       = db.Column(db.String(20), nullable=True)
    sbe_manual_override = db.Column(db.Boolean, default=False)
    
    cert_status       = db.Column(db.String(20), default="Pendiente") 
    cert_fecha        = db.Column(db.Date, nullable=True)
    
    final_remito      = db.Column(db.String(50), nullable=True)
    final_peso        = db.Column(db.Float, nullable=True)
    observation_reason = db.Column(db.String(100), nullable=True)

    frozen_flete_price = db.Column(db.Float, nullable=True)
    frozen_arena_price = db.Column(db.Float, nullable=True)
    frozen_merma_money = db.Column(db.Float, default=0.0) 
    frozen_flete_neto  = db.Column(db.Float, default=0.0)
    frozen_flete_iva   = db.Column(db.Float, default=0.0)

class Quota(db.Model):
    __tablename__ = "quota"
    id               = db.Column(db.Integer, primary_key=True)
    transportista_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    arenera_id       = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    date             = db.Column(db.Date, nullable=False, index=True)
    limit            = db.Column(db.Integer, nullable=False, default=0)
    used             = db.Column(db.Integer, nullable=False, default=0)
    arenera          = db.relationship("User", foreign_keys=[arenera_id])
    __table_args__   = (
        db.UniqueConstraint("transportista_id", "arenera_id", "date", name="uix_quota"),
    )

class Chofer(db.Model):
    __tablename__ = "chofer"
    id                  = db.Column(db.Integer, primary_key=True)
    transportista_id    = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    nombre              = db.Column(db.String(100), nullable=False)
    dni                 = db.Column(db.String(50), unique=True, nullable=False, index=True)
    gender              = db.Column(db.String(10), nullable=True)
    tractor             = db.Column(db.String(20), nullable=True)
    trailer             = db.Column(db.String(20), nullable=True)
    tipo                = db.Column(db.String(20), nullable=True)
    transportista       = db.relationship("User", backref=db.backref("choferes", lazy="dynamic"))
    __table_args__      = (
        db.UniqueConstraint("transportista_id", "dni", name="uix_transportista_dni"),
    )

class SystemConfig(db.Model):
    __tablename__ = "system_config"
    id             = db.Column(db.Integer, primary_key=True)
    tolerance_kg   = db.Column(db.Float, default=700.0)
    dispatch_price = db.Column(db.Float, default=0.0)
    sand_price     = db.Column(db.Float, default=0.0)
    transport_price= db.Column(db.Float, default=0.0)
    admin_email    = db.Column(db.String(120), nullable=True)

# ----------------------------
# Bootstrapping DB
# ----------------------------
with app.app_context():
    if DB_SCHEMA and DB_SCHEMA != "public":
        try:
            db.session.execute(text(f"CREATE SCHEMA IF NOT EXISTS {DB_SCHEMA}"))
            db.session.commit()
        except Exception:
            db.session.rollback()

    db.create_all()

    try:
        db.session.execute(text('ALTER TABLE "user" ALTER COLUMN password_hash TYPE VARCHAR(512)'))
        db.session.commit()
    except Exception:
        db.session.rollback()

    admin = db.session.query(User).filter(func.lower(User.username) == norm_username(ADMIN_USER)).first()
    if not admin:
        admin = User(
            username=norm_username(ADMIN_USER),
            password_hash=generate_password_hash(ADMIN_PASS),
            tipo="admin",
        )
        db.session.add(admin)
        db.session.commit()

# ----------------------------
# Decoradores de auth
# ----------------------------
def login_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return wrapped

def role_required(*roles):
    def deco(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if session.get("tipo") not in roles:
                flash("No tienes permisos para acceder a esta sección.", "error")
                return redirect(url_for("login_page"))
            return f(*args, **kwargs)
        return wrapped
    return deco

# ----------------------------
# LOGIN / LOGOUT
# ----------------------------
@app.route("/")
def login_page():
    return render_template(tpl("login"))

@app.route("/login", methods=["POST"])
def login():
    u = norm_username(request.form.get("usuario"))
    p = (request.form.get("clave") or "").strip()
    
    user = db.session.query(User).filter(func.lower(User.username) == u).first()
    
    if user and check_password_hash(user.password_hash, p):
        session.clear()
        session["user_id"] = user.id
        session["tipo"]    = user.tipo
        session.modified = True 

        if user.tipo == "admin":
            return redirect(url_for("admin_panel"))
        if user.tipo == "gestion":
            return redirect(url_for("admin_dashboard"))
        if user.tipo == "transportista":
            return redirect(url_for("transportista_panel"))
        if user.tipo == "arenera":
            return redirect(url_for("arenera_panel"))
            
    flash("Usuario o clave incorrectos", "error")
    return redirect(url_for("login_page"))

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login_page"))

@app.route("/change_password", methods=["GET", "POST"])
@login_required
def change_password():
    user = User.query.get(session["user_id"])
    if request.method == "POST":
        actual = request.form["current_password"].strip()
        nueva  = request.form["new_password"].strip()
        if check_password_hash(user.password_hash, actual):
            user.password_hash = generate_password_hash(nueva, method="pbkdf2:sha256")
            db.session.commit()
            flash("Contraseña actualizada correctamente.", "success")
            return redirect(url_for(f"{user.tipo}_panel"))
        flash("Contraseña actual no coincide", "error")
    return render_template(tpl("change_password"), user=user)

# ----------------------------
# PANEL ADMIN
# ----------------------------
@app.route("/admin")
@login_required
@role_required("admin")
def admin_panel():
    today = get_arg_today()
    week_start = today - timedelta(days=today.weekday())
    week_end   = week_start + timedelta(days=6)
    
    stats_shipments = db.session.query(
        Shipment.transportista_id, 
        func.count(Shipment.id)
    ).group_by(Shipment.transportista_id).all()
    map_shipments = {tid: count for tid, count in stats_shipments}

    stats_quotas = db.session.query(
        Quota.transportista_id,
        func.sum(Quota.limit).label('total'),
        func.sum(Quota.used).label('used')
    ).filter(
        Quota.date >= week_start,
        Quota.date <= week_end
    ).group_by(Quota.transportista_id).all()
    map_quotas = {tid: (limit or 0, used or 0) for tid, limit, used in stats_quotas}

    all_transportistas = User.query.filter_by(tipo="transportista").all()
    t_list = []
    
    for t in all_transportistas:
        total_sent = map_shipments.get(t.id, 0)
        limit, used = map_quotas.get(t.id, (0, 0))
        remaining = max(0, limit - used)

        t_list.append({
            "t": t, 
            "sent": total_sent, 
            "quota": remaining,
            "quota_limit": limit,
            "quota_used": used
        })

    stats_areneras = db.session.query(
        Shipment.arenera_id, func.count(Shipment.id)
    ).group_by(Shipment.arenera_id).all()
    map_areneras_ship = {aid: count for aid, count in stats_areneras}

    a_list = []
    for a in User.query.filter_by(tipo="arenera", parent_id=None).all():
        total_ship = map_areneras_ship.get(a.id, 0)
        a_list.append({"a": a, "shipments": total_ship})

    sub_a_list = User.query.filter(User.tipo=="arenera", User.parent_id != None).all()
    g_list = User.query.filter_by(tipo="gestion").all()

    return render_template(
        tpl("admin_panel"),
        transportistas=t_list,
        areneras=a_list,
        sub_areneras=sub_a_list,
        gestores=g_list,      
    )

@app.route("/admin/create_user", methods=["POST"])
@login_required
@role_required("admin")
def create_user():
    uname = norm_username(request.form.get("username"))
    pwd   = (request.form.get("password") or "").strip()
    tipo  = request.form.get("tipo")
    
    parent_id = None
    if tipo == "sub_arenera":
        tipo = "arenera"
        try:
            parent_id = int(request.form.get("parent_id"))
        except (TypeError, ValueError):
            flash("Debes seleccionar una Arenera Padre para el sub-usuario.", "error")
            return redirect(url_for("admin_panel"))

    if not uname or not pwd or tipo not in ("transportista", "arenera", "gestion", "admin"):
        flash("Datos inválidos", "error")
        return redirect(url_for("admin_panel"))

    if db.session.query(User).filter(func.lower(User.username) == uname).first():
        flash("Usuario ya existe.", "error")
        return redirect(url_for("admin_panel"))

    nuevo = User(
        username=uname, 
        password_hash=generate_password_hash(pwd), 
        tipo=tipo,
        parent_id=parent_id
    )
    db.session.add(nuevo)
    db.session.commit()

    flash("Usuario creado.", "success")
    return redirect(url_for("admin_panel"))

@app.post("/admin/reset_password_admin")
@login_required
@role_required("admin")
def reset_password_admin():
    user_id = request.form.get("user_id")
    new_pass = request.form.get("new_password")
    
    if not user_id or not new_pass:
        flash("Faltan datos.", "error")
        return redirect(url_for("admin_panel"))
        
    u = db.session.get(User, user_id)
    if not u:
        flash("Usuario no encontrado.", "error")
        return redirect(url_for("admin_panel"))
        
    u.password_hash = generate_password_hash(new_pass)
    db.session.commit()
    
    flash(f"Clave actualizada para {u.username}. Nueva clave temporal: {new_pass}", "success")
    return redirect(url_for("admin_panel"))

@app.route("/admin/delete_user/<int:user_id>")
@login_required
@role_required("admin")
def delete_user(user_id):
    u = User.query.get_or_404(user_id)

    if u.tipo == "arenera":
        has_shipments = Shipment.query.filter_by(arenera_id=u.id).count() > 0
        has_quotas    = Quota.query.filter_by(arenera_id=u.id).count() > 0
        if has_shipments or has_quotas:
            flash("No se puede eliminar la arenera porque tiene datos asociados (viajes/cuotas).", "error")
            return redirect(url_for("admin_panel"))

    db.session.delete(u)
    db.session.commit()
    flash("Usuario eliminado", "success")
    return redirect(url_for("admin_panel"))

@app.route("/admin/quotas/<int:transportista_id>", methods=["GET", "POST"])
@login_required
@role_required("admin")
def manage_quotas(transportista_id):
    t = User.query.get_or_404(transportista_id)
    areneras = User.query.filter(User.tipo == "arenera", User.parent_id == None).order_by(User.username.asc()).all()

    if request.method == "POST":
        start_str = (request.form.get("start_date") or "").strip()
    else:
        start_str = (request.args.get("start") or "").strip()

    try:
        start_date = date.fromisoformat(start_str) if start_str else get_arg_today()
    except ValueError:
        start_date = get_arg_today()

    week_dates = [start_date + timedelta(days=i) for i in range(7)]

    if request.method == "POST":
        total_inserts = total_updates = total_deletes = 0

        try:
            for a in areneras:
                for d in week_dates:
                    key = f"q_{a.id}_{d.isoformat()}"
                    raw = (request.form.get(key) or "").strip()

                    if raw == "" or raw == "0":
                        q = db.session.query(Quota).filter_by(
                            transportista_id=transportista_id,
                            arenera_id=a.id,
                            date=d
                        ).first()
                        if q:
                            db.session.delete(q)
                            total_deletes += 1
                        continue

                    try:
                        v = int(raw)
                    except ValueError:
                        continue
                    if v < 0: v = 0

                    q = db.session.query(Quota).filter_by(
                        transportista_id=transportista_id,
                        arenera_id=a.id,
                        date=d
                    ).first()

                    if q:
                        q.limit = v
                        if (q.used or 0) > v:
                            q.used = v
                        total_updates += 1
                    else:
                        q = Quota(
                            transportista_id=transportista_id,
                            arenera_id=a.id,
                            date=d,
                            limit=v,
                            used=0
                        )
                        db.session.add(q)
                        total_inserts += 1

            db.session.commit()
            flash(f"Cuotas guardadas: {total_inserts} nuevas, {total_updates} actualizadas.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Error al guardar cuotas: {e}", "error")

        return redirect(url_for("manage_quotas", transportista_id=transportista_id, start=start_date.isoformat()))

    existing = (
        Quota.query
        .filter(
            Quota.transportista_id == transportista_id,
            Quota.date.in_(week_dates)
        ).all()
    )
    quota_map = {(q.arenera_id, q.date): q for q in existing}

    from sqlalchemy import func
    used_rows = (
        db.session.query(
            Shipment.arenera_id,
            Shipment.date,
            func.count(Shipment.id)
        )
        .filter(
            Shipment.transportista_id == transportista_id,
            Shipment.date >= week_dates[0],
            Shipment.date <= week_dates[-1],
        )
        .group_by(Shipment.arenera_id, Shipment.date)
        .all()
    )
    used_map = {(aid, d): int(c) for (aid, d, c) in used_rows}

    return render_template(
        tpl("admin_quotas"),
        transportista=t,
        areneras=areneras,
        week_dates=week_dates,
        quota_map=quota_map,
        used_map=used_map,
        start_date=start_date
    )

@app.route("/admin/todos_camiones")
@login_required
@role_required("admin")
def todos_camiones():
    status        = request.args.get("status")
    transportista = (request.args.get("transportista") or "").strip().lower()
    arenera       = (request.args.get("arenera") or "").strip().lower()
    dfrom_str     = (request.args.get("from") or "").strip()
    dto_str       = (request.args.get("to") or "").strip()

    q = Shipment.query

    try:
        if dfrom_str:
            q = q.filter(Shipment.date >= date.fromisoformat(dfrom_str))
        if dto_str:
            q = q.filter(Shipment.date <= date.fromisoformat(dto_str))
    except ValueError:
        pass

    if status:
        q = q.filter(Shipment.status == status)

    shipments = q.order_by(Shipment.date.desc(), Shipment.id.desc()).all()

    if transportista:
        shipments = [s for s in shipments if transportista in s.transportista.username.lower()]
    if arenera:
        shipments = [s for s in shipments if arenera in s.arenera.username.lower()]

    return render_template(
        tpl("todos_camiones"),
        shipments=shipments, status=status, transportista=transportista, arenera=arenera,
        dfrom=dfrom_str, dto=dto_str
    )

# --- Asegúrate de tener estos imports al inicio de app.py o dentro de la función ---
from sqlalchemy import cast, Date

@app.post("/admin/dashboard_data")
@login_required
@role_required("admin", "gestion")
@csrf.exempt 
def admin_dashboard_data():
    # --- A. FILTROS Y CONSULTA GENERAL ---
    data = request.get_json(silent=True) or {}
    dfrom, dto = _resolve_range(data)
    arenera_id = data.get("arenera_id")
    
    # 1. QUERY GENERAL (Para KPIs, Listas y Tablas)
    # Mantenemos esta lógica para listar los viajes iniciados en el periodo
    base_q = db.session.query(Shipment).filter(Shipment.date >= dfrom, Shipment.date <= dto)
    
    if arenera_id and str(arenera_id) != "all":
        base_q = base_q.filter(Shipment.arenera_id == int(arenera_id))
    
    all_ships = base_q.all()

    # --- B. LÓGICA DE CHART (GRÁFICO CORREGIDO) ---
    # En lugar de usar 'all_ships', hacemos dos consultas independientes para llenar el gráfico correctamente.
    
    # Inicializar el eje X (Fechas)
    hoy_arg = get_arg_today()
    fecha_corte = min(dto, hoy_arg)
    delta_days = (fecha_corte - dfrom).days
    if delta_days < 0: delta_days = -1

    daily_stats = {}
    for i in range(delta_days + 1):
        day_loop = dfrom + timedelta(days=i)
        daily_stats[day_loop.isoformat()] = {"out": 0.0, "in": 0.0}

    # 1. Consulta DESPACHADO (Eje fecha = Fecha Salida)
    q_out = db.session.query(
        Shipment.date, 
        func.sum(Shipment.peso_neto_arenera)
    ).filter(
        Shipment.date >= dfrom, 
        Shipment.date <= fecha_corte
    )
    if arenera_id and str(arenera_id) != "all":
        q_out = q_out.filter(Shipment.arenera_id == int(arenera_id))
    
    results_out = q_out.group_by(Shipment.date).all()
    
    for r_date, r_sum in results_out:
        d_str = r_date.isoformat()
        if d_str in daily_stats:
            daily_stats[d_str]["out"] = float(r_sum or 0)

    # 2. Consulta RECIBIDO (Eje fecha = Fecha LLEGADA SBE) [CORRECCIÓN AQUI]
    # Aquí filtramos por fecha de llegada, independiente de cuándo salió
    q_in = db.session.query(
        cast(Shipment.sbe_fecha_llegada, Date), 
        func.sum(Shipment.sbe_peso_neto) # Usamos el peso SBE
    ).filter(
        cast(Shipment.sbe_fecha_llegada, Date) >= dfrom,
        cast(Shipment.sbe_fecha_llegada, Date) <= fecha_corte
    )
    if arenera_id and str(arenera_id) != "all":
        q_in = q_in.filter(Shipment.arenera_id == int(arenera_id))

    results_in = q_in.group_by(cast(Shipment.sbe_fecha_llegada, Date)).all()

    for r_date, r_sum in results_in:
        if r_date: # Puede ser None si hay error de datos
            d_str = r_date.isoformat()
            if d_str in daily_stats:
                daily_stats[d_str]["in"] = float(r_sum or 0)


    # --- C. RESTO DE LA LÓGICA (KPIs, Listas, etc.) ---
    # Usamos all_ships (Base de Salidas) para el resto de las estadísticas
    
    conf = get_config()
    tol_tn = conf.tolerance_kg / 1000.0

    ships_to_source = [s for s in all_ships if (not s.peso_neto_arenera or s.peso_neto_arenera == 0) and not s.sbe_fecha_llegada]
    ships_to_dest = [s for s in all_ships if (s.peso_neto_arenera and s.peso_neto_arenera > 0) and not s.sbe_fecha_llegada]
    ships_arrived = [s for s in all_ships if s.sbe_fecha_llegada is not None]
    
    count_to_source = len(ships_to_source)
    count_to_dest = len(ships_to_dest)
    tn_to_dest    = sum(s.peso_neto_arenera or 0 for s in ships_to_dest)
    count_arrived = len(ships_arrived)
    
    # OJO: Para el KPI de "Tn Llegadas" en la tarjeta, es mejor usar la suma real del gráfico (lo que llegó en este periodo)
    # en lugar de "lo que llegó de los camiones que salieron en este periodo".
    # Sumamos todo lo del gráfico IN:
    tn_arrived_periodo = sum(v["in"] for v in daily_stats.values())

    total_viajes = len(all_ships)
    total_moving = count_to_source + count_to_dest
    tn_despachadas_total = sum(s.peso_neto_arenera or 0 for s in all_ships)

    payments_detail = []
    arenera_volumen = {}
    top_trans_map = {}
    total_costo_proyectado = 0.0

    for s in all_ships:
        # Rankings (Usamos base salida)
        a_name = s.arenera.username
        arenera_volumen[a_name] = arenera_volumen.get(a_name, 0) + (s.peso_neto_arenera or 0)
        
        t_name = s.transportista.username
        w_trans = s.final_peso or s.sbe_peso_neto or 0
        top_trans_map[t_name] = top_trans_map.get(t_name, 0) + w_trans

        # Cálculos Dinero (Igual que antes)
        neto_flete = 0.0
        neto_arena = 0.0
        
        val_w_in = s.final_peso or s.sbe_peso_neto or 0
        val_w_out = s.peso_neto_arenera or 0

        if val_w_in > 0:
            if s.frozen_flete_neto is not None:
                neto_flete = s.frozen_flete_neto
            else:
                price_flete = s.transportista.custom_price or 0
                price_arena_ref = s.arenera.custom_price or 0
                tn_base = val_w_out if s.arenera.cert_type == 'salida' else val_w_in
                merma_money = 0.0
                if s.arenera.cert_type != 'salida':
                    diff = val_w_out - val_w_in
                    if diff > tol_tn: merma_money = (diff - tol_tn) * price_arena_ref
                neto_flete = (tn_base * price_flete) - merma_money
        
        iva_flete = max(0, neto_flete * 1.21)
        total_costo_proyectado += iva_flete

        if val_w_out > 0:
            p_arena = s.frozen_arena_price if s.frozen_arena_price is not None else (s.arenera.custom_price or 0)
            neto_arena = val_w_out * p_arena
        
        iva_arena = max(0, neto_arena * 1.21)
        total_costo_proyectado += iva_arena

        if s.cert_status == "Certificado" and s.cert_fecha:
            if iva_flete > 0:
                d_pay = s.cert_fecha + timedelta(days=(s.transportista.payment_days or 30))
                payments_detail.append({
                    "raw": d_pay, "fecha": d_pay.strftime("%d/%m/%Y"),
                    "monto": iva_flete, "entidad": s.transportista.username, "tipo": "Flete",
                    "dia_semana": ["Lun","Mar","Mié","Jue","Vie","Sáb","Dom"][d_pay.weekday()]
                })
            if iva_arena > 0:
                d_pay = s.cert_fecha + timedelta(days=(s.arenera.payment_days or 30))
                payments_detail.append({
                    "raw": d_pay, "fecha": d_pay.strftime("%d/%m/%Y"),
                    "monto": iva_arena, "entidad": s.arenera.username, "tipo": "Arena",
                    "dia_semana": ["Lun","Mar","Mié","Jue","Vie","Sáb","Dom"][d_pay.weekday()]
                })

    # --- D. PREPARAR JSON FINAL ---
    sorted_days = sorted(daily_stats.keys())
    chart_labels = [datetime.strptime(d, "%Y-%m-%d").strftime("%d/%m") for d in sorted_days]
    data_out = [daily_stats[d]["out"] for d in sorted_days]
    data_in  = [daily_stats[d]["in"]  for d in sorted_days]

    pay_map = {}
    for p in payments_detail:
        k = f"{p['raw']}_{p['entidad']}_{p['tipo']}"
        if k not in pay_map: pay_map[k] = {**p, "count": 0, "monto": 0}
        pay_map[k]["monto"] += p["monto"]
        pay_map[k]["count"] += 1
    
    payments_final = sorted(pay_map.values(), key=lambda x: x["raw"])
    sorted_areneras = sorted(arenera_volumen.items(), key=lambda x: x[1], reverse=True)
    sorted_trans = sorted(top_trans_map.items(), key=lambda x: x[1], reverse=True)[:5]

    return {
        "kpi": {
            "viajes_total": total_viajes,
            "viajes_ruta": total_moving,
            "to_source": count_to_source,
            "to_dest": count_to_dest,
            "tn_to_dest": tn_to_dest,
            "arrived": count_arrived,
            "tn_out": tn_despachadas_total,
            "tn_in": tn_arrived_periodo, # USAMOS EL VALOR REAL DE LLEGADA
            "costo_total": total_costo_proyectado
        },
        "chart_data": { "labels": chart_labels, "out": data_out, "in": data_in },
        "payments_table": payments_final,
        "arenera_stats": [{"name": k, "value": v} for k, v in sorted_areneras],
        "top_trans": {
            "labels": [x[0] for x in sorted_trans],
            "values": [x[1] for x in sorted_trans]
        }
    }

@app.route("/admin/dashboard")
@login_required
@role_required("admin", "gestion")
def admin_dashboard():
    areneras = User.query.filter_by(tipo="arenera", parent_id=None).order_by(User.username).all()
    return render_template(tpl("admin_dashboard"), areneras=areneras)

# Asegúrate de tener estos imports al principio del archivo
from sqlalchemy import cast, Date, or_

# Asegúrate de tener: from sqlalchemy import cast, Date, or_

@app.route("/admin/certificacion")
@login_required
@role_required("admin", "gestion")
def admin_certificacion():
    # --- Parámetros existentes ---
    view_mode  = request.args.get("view", "pendiente")
    tid_filter = request.args.get("transportista_id")
    aid_filter = request.args.get("arenera_id")
    status_filter = request.args.get("status")
    
    # Filtros de fecha de VIAJE (Salida)
    start_str = request.args.get("start")
    end_str   = request.args.get("end")
    
    # --- Parámetros de Búsqueda y RANGO SBE ---
    search_q = (request.args.get("search") or "").strip()
    
    # [NUEVO] Capturamos inicio y fin de llegada SBE
    arr_start_str = (request.args.get("arrival_start") or "").strip()
    arr_end_str   = (request.args.get("arrival_end") or "").strip()

    # --- Lógica de Fechas Viaje (Existente) ---
    use_date_filter = False
    if start_str and end_str:
        try:
            start_date = date.fromisoformat(start_str)
            end_date   = date.fromisoformat(end_str)
            use_date_filter = True
        except ValueError:
            pass

    # --- Listas ---
    trans_list = User.query.filter_by(tipo="transportista").order_by(User.username).all()
    aren_list  = User.query.filter_by(tipo="arenera", parent_id=None).order_by(User.username).all()

    q = Shipment.query
    
    # --- Filtros Base ---
    if view_mode == "historial":
        q = q.filter(Shipment.cert_status == "Certificado")
        if use_date_filter:
            q = q.filter(Shipment.cert_fecha >= start_date, Shipment.cert_fecha <= end_date)
        q = q.order_by(Shipment.cert_fecha.desc(), Shipment.id.desc())
    else:
        q = q.filter(Shipment.cert_status != "Certificado")
        if use_date_filter:
            q = q.filter(Shipment.date >= start_date, Shipment.date <= end_date)
        q = q.order_by(case((Shipment.cert_status == "Observado", 1), else_=2), Shipment.date.desc())

    # --- Aplicar Filtros Dropdown ---
    if tid_filter and tid_filter != "all":
        q = q.filter(Shipment.transportista_id == int(tid_filter))
    if aid_filter and aid_filter != "all":
        q = q.filter(Shipment.arenera_id == int(aid_filter))
    if status_filter and status_filter != "all":
        q = q.filter(Shipment.cert_status == status_filter)

    # --- Filtro de Búsqueda ---
    if search_q:
        q = q.filter(
            or_(
                Shipment.remito_arenera.ilike(f"%{search_q}%"),
                Shipment.sbe_remito.ilike(f"%{search_q}%"),
                Shipment.tractor.ilike(f"%{search_q}%")
            )
        )

    # --- [NUEVO] FILTRO RANGO LLEGADA SBE ---
    if arr_start_str:
        try:
            d_start = date.fromisoformat(arr_start_str)
            # Usamos cast(..Date) para ignorar la hora y comparar solo el día
            q = q.filter(cast(Shipment.sbe_fecha_llegada, Date) >= d_start)
        except ValueError: pass

    if arr_end_str:
        try:
            d_end = date.fromisoformat(arr_end_str)
            q = q.filter(cast(Shipment.sbe_fecha_llegada, Date) <= d_end)
        except ValueError: pass

    shipments = q.all()
    
    return render_template(
        tpl("admin_certification"), 
        shipments=shipments, 
        view_mode=view_mode,
        transportistas=trans_list, 
        areneras=aren_list,
        sel_tid=tid_filter,        
        sel_aid=aid_filter,
        sel_status=status_filter,
        sel_search=search_q,       
        
        # Pasamos las nuevas variables al template
        sel_arr_start=arr_start_str,
        sel_arr_end=arr_end_str,
        
        start_date=start_str, 
        end_date=end_str
    )

@app.post("/admin/certificar/<int:shipment_id>")
@login_required
@role_required("admin")
def certify_shipment(shipment_id):
    s = Shipment.query.get_or_404(shipment_id)
    
    s.final_remito = s.sbe_remito if s.sbe_remito else s.remito_arenera
    peso_final = s.sbe_peso_neto if (s.sbe_peso_neto and s.sbe_peso_neto > 0) else s.peso_neto_arenera
    s.final_peso = peso_final

    price_flete = s.transportista.custom_price or 0
    price_arena = s.arenera.custom_price or 0
    
    conf = get_config()
    tol_tn = conf.tolerance_kg / 1000.0
    
    peso_salida = s.peso_neto_arenera or 0
    merma_money = 0.0
    tn_base_flete = 0.0

    if s.arenera.cert_type == 'salida':
        tn_base_flete = peso_salida
        merma_money = 0.0
    else:
        tn_base_flete = peso_final
        diff = peso_salida - peso_final
        if diff > tol_tn:
            excess = diff - tol_tn
            merma_money = excess * price_arena

    flete_neto = (tn_base_flete * price_flete) - merma_money
    flete_iva  = flete_neto * 1.21
    if flete_iva < 0: flete_iva = 0 

    s.frozen_flete_price = price_flete
    s.frozen_arena_price = price_arena
    s.frozen_merma_money = merma_money
    s.frozen_flete_neto  = flete_neto
    s.frozen_flete_iva   = flete_iva

    s.cert_status = "Certificado"
    s.cert_fecha  = get_arg_today()
    if s.status != "Llego": 
        s.status = "Llego"

    db.session.commit()
    flash(f"Viaje #{s.id} certificado. Neto: ${flete_neto:,.0f} (Multa: ${merma_money:,.0f})", "success")
    return redirect(url_for("admin_certificacion"))

@app.post("/admin/corregir_sbe/<int:shipment_id>")
@login_required
@role_required("admin")
def corregir_sbe(shipment_id):
    s = Shipment.query.get_or_404(shipment_id)
    
    if request.form.get("action") == "reset":
        s.sbe_remito = None
        s.sbe_peso_neto = None
        s.sbe_fecha_salida = None
        s.sbe_fecha_llegada = None
        s.sbe_patente = None
        s.sbe_manual_override = False
        s.cert_status = "Pendiente"
        s.observation_reason = None
        db.session.commit()
        flash("Corrección manual eliminada. Listo para re-sincronizar.", "info")
        return redirect(url_for("admin_certificacion"))

    nuevo_remito = request.form.get("sbe_remito", "").strip()
    nuevo_peso   = request.form.get("sbe_peso", "").strip()
    
    if nuevo_remito: 
        s.sbe_remito = nuevo_remito
    if nuevo_peso:   
        try:
            s.sbe_peso_neto = float(nuevo_peso.replace(",", "."))
        except ValueError:
            flash("Error: Peso inválido.", "error")
            return redirect(url_for("admin_certificacion"))
    
    s.sbe_manual_override = True 
    if not s.sbe_fecha_salida:
        s.sbe_fecha_salida = datetime.combine(s.date, datetime.min.time())
    if not s.sbe_fecha_llegada:
        s.sbe_fecha_llegada = s.sbe_fecha_salida

    if s.sbe_remito and s.sbe_peso_neto is not None:
         s.cert_status = "Pre-Aprobado"
         s.observation_reason = "Corregido Manualmente"
         
    db.session.commit()
    flash("Datos corregidos. Fechas ajustadas automáticamente.", "success")
    return redirect(url_for("admin_certificacion"))

@app.post("/admin/certificar_masivo")
@login_required
@role_required("admin")
def certify_batch():
    # 1. Recuperar TODOS los Filtros del formulario
    tid_filter = request.form.get("transportista_id")
    aid_filter = request.form.get("arenera_id")
    start_str  = request.form.get("start")
    end_str    = request.form.get("end")
    
    # [NUEVOS FILTROS]
    search_q      = request.form.get("search")
    arr_start_str = request.form.get("arrival_start")
    arr_end_str   = request.form.get("arrival_end")
    
    # 2. Construir la Query Base (Solo Pre-Aprobados)
    q = Shipment.query.filter(Shipment.cert_status == "Pre-Aprobado")

    # --- APLICAR FILTROS (Idéntico a la vista principal) ---

    # A. Fechas de Salida (Viaje)
    if start_str and end_str:
        try:
            s_date = date.fromisoformat(start_str)
            e_date = date.fromisoformat(end_str)
            q = q.filter(Shipment.date >= s_date, Shipment.date <= e_date)
        except ValueError: pass

    # B. Empresas
    if tid_filter and tid_filter != "all":
        q = q.filter(Shipment.transportista_id == int(tid_filter))
    if aid_filter and aid_filter != "all":
        q = q.filter(Shipment.arenera_id == int(aid_filter))

    # C. Búsqueda Texto
    if search_q:
        q = q.filter(
            or_(
                Shipment.remito_arenera.ilike(f"%{search_q}%"),
                Shipment.sbe_remito.ilike(f"%{search_q}%"),
                Shipment.tractor.ilike(f"%{search_q}%")
            )
        )

    # D. Fechas de Llegada SBE
    if arr_start_str:
        try:
            d_start = date.fromisoformat(arr_start_str)
            q = q.filter(cast(Shipment.sbe_fecha_llegada, Date) >= d_start)
        except ValueError: pass

    if arr_end_str:
        try:
            d_end = date.fromisoformat(arr_end_str)
            q = q.filter(cast(Shipment.sbe_fecha_llegada, Date) <= d_end)
        except ValueError: pass

    targets = q.all()
    
    if not targets:
        flash("No hay viajes 'Pre-Aprobados' que coincidan con los filtros para certificar.", "info")
        return redirect(url_for('admin_certificacion', **request.args))

    # 3. Procesamiento Masivo
    count = 0
    conf = get_config()
    tol_tn = conf.tolerance_kg / 1000.0
    hoy = get_arg_today()

    for s in targets:
        # Lógica de certificación (Idéntica a la individual)
        s.final_remito = s.sbe_remito
        s.final_peso   = s.sbe_peso_neto
        
        price_flete = s.transportista.custom_price or 0
        price_arena = s.arenera.custom_price or 0
        
        peso_salida = s.peso_neto_arenera or 0
        peso_final  = s.final_peso or 0
        
        merma_money = 0.0
        tn_base_flete = 0.0

        if s.arenera.cert_type == 'salida':
            tn_base_flete = peso_salida
        else:
            tn_base_flete = peso_final
            diff = peso_salida - peso_final
            if diff > tol_tn:
                excess = diff - tol_tn
                merma_money = excess * price_arena

        flete_neto = (tn_base_flete * price_flete) - merma_money
        flete_iva  = max(0, flete_neto * 1.21)

        s.frozen_flete_price = price_flete
        s.frozen_arena_price = price_arena
        s.frozen_merma_money = merma_money
        s.frozen_flete_neto  = flete_neto
        s.frozen_flete_iva   = flete_iva

        s.cert_status = "Certificado"
        s.cert_fecha  = hoy
        if s.status != "Llego": 
            s.status = "Llego"
            
        count += 1

    db.session.commit()
    flash(f"✅ Éxito: Se certificaron masivamente {count} viajes.", "success")
    
    # Redirigir manteniendo los filtros visuales (Reconstruimos la URL)
    return redirect(url_for('admin_certificacion', 
                            transportista_id=tid_filter, 
                            arenera_id=aid_filter, 
                            start=start_str, 
                            end=end_str,
                            search=search_q,
                            arrival_start=arr_start_str,
                            arrival_end=arr_end_str,
                            view='pendiente'))

@app.route("/transportista/panel")
@login_required
@role_required("transportista")
def transportista_panel():
    u   = db.session.get(User, session["user_id"])
    hoy = get_arg_today()

    week_start = hoy - timedelta(days=hoy.weekday())
    week_end   = week_start + timedelta(days=6)
    next_monday = week_start + timedelta(days=7)

    # CORRECCIÓN KPI: Sumamos "En viaje", "Salió" y "Salido a SBE"
    # Básicamente todo lo que NO sea "Llego" ni esté Certificado.
    en_viaje_total = (Shipment.query
        .filter(
            Shipment.transportista_id == u.id,
            Shipment.status.in_(["En viaje", "Salió", "Salido a SBE", "En Viaje"]) # Agregamos variantes para asegurar
        )
        .count())

    llegados_semana = (Shipment.query
        .filter(
            Shipment.transportista_id == u.id,
            Shipment.status == "Llego", # O status que signifique finalizado
            Shipment.date >= week_start,
            Shipment.date <  next_monday
        ).count())
    
    choferes_list = Chofer.query.filter_by(transportista_id=u.id).all() 

    # Cálculo de Cupos (Igual que antes)
    q_stats = db.session.query(
        func.sum(Quota.limit).label('total'),
        func.sum(Quota.used).label('used')
    ).filter(
        Quota.transportista_id == u.id,
        Quota.date >= week_start,
        Quota.date <= week_end
    ).first()

    q_limit = q_stats.total or 0
    q_used  = q_stats.used or 0
    q_remaining = max(0, q_limit - q_used)

    stats = {
        "en_viaje":   en_viaje_total,
        "llegados":   llegados_semana,
        "week_from":  week_start,
        "week_to":    week_end,
        "link_choferes": url_for("transportista_choferes"),
        "cupo_restante": q_remaining,
        "cupo_total": q_limit
    }

    envios = (Shipment.query
              .filter_by(transportista_id=u.id)
              .order_by(Shipment.date.desc(), Shipment.id.desc())
              .all())

    assignments = []
    for q in Quota.query.filter_by(transportista_id=u.id, date=hoy).all():
        if q.arenera:
            assignments.append({
                "arenera": q.arenera,
                "daily_limit": q.limit,
                "daily_used": q.used
            })

    return render_template(
        tpl("transportista_panel"),
        user=u, today=hoy,
        stats=stats,
        assignments=assignments,
        envios=envios,
        choferes=choferes_list, 
    )

@app.route("/transportista/history")
@login_required
@role_required("transportista")
def transportista_history():
    u = db.session.get(User, session["user_id"])

    start_str = (request.args.get("start_date") or "").strip()
    end_str   = (request.args.get("end_date") or "").strip()
    status    = (request.args.get("status") or "").strip()
    search    = (request.args.get("search") or "").strip().lower()

    min_date = db.session.query(func.min(Shipment.date)).filter_by(transportista_id=u.id).scalar() or get_arg_today()
    max_date = db.session.query(func.max(Shipment.date)).filter_by(transportista_id=u.id).scalar() or get_arg_today()

    try:
        start_date = date.fromisoformat(start_str) if start_str else min_date
    except ValueError:
        start_date = min_date
    try:
        end_date = date.fromisoformat(end_str) if end_str else max_date
    except ValueError:
        end_date = max_date
    if end_date < start_date:
        start_date, end_date = end_date, start_date

    base_rows = (Shipment.query
                 .filter(
                     Shipment.transportista_id == u.id,
                     Shipment.date >= start_date,
                     Shipment.date <= end_date,
                 )
                 .order_by(Shipment.date.desc(), Shipment.id.desc())
                 .all())

    if search:
        search_id = -1
        clean_search = search.replace("#", "")
        if clean_search.isdigit():
            search_id = int(clean_search)
        
        needle = search
        def hit(s: Shipment) -> bool:
            if s.id == search_id: return True
            vals = [
                s.chofer, s.dni, s.tractor, s.trailer, s.tipo,
                getattr(s.arenera, "username", ""),
                s.remito_arenera, 
                s.final_remito
            ]
            return any(needle in (v or "").lower() for v in vals)
        base_rows = [s for s in base_rows if hit(s)]

    counts = {
        "total":      len(base_rows),
        "en_viaje":   sum(1 for s in base_rows if s.status == "En viaje"),
        "llegados":   sum(1 for s in base_rows if s.status == "Llego"),
    }

    if status:
        rows = [s for s in base_rows if s.status == status]
    else:
        rows = base_rows

    return render_template(
        tpl("transportista_history"),
        user=u,
        shipments=rows,
        counts=counts,
        start_date=start_date.isoformat(),
        end_date=end_date.isoformat(),
        status=status,
        search=(request.args.get("search") or "").strip(),
    )

@app.get("/transportista/export")
@login_required
@role_required("transportista")
def transportista_export():
    u = db.session.get(User, session["user_id"])
    today = get_arg_today()
    start_str = request.args.get("start")
    end_str   = request.args.get("end")

    try:
        if start_str:
            start_date = date.fromisoformat(start_str)
        else:
            start_date = today.replace(day=1)
        if end_str:
            end_date = date.fromisoformat(end_str)
        else:
            end_date = today
    except ValueError:
        start_date = today.replace(day=1)
        end_date   = today

    rows = (db.session.query(Shipment)
            .outerjoin(User, Shipment.operador_id == User.id)
            .filter(
                Shipment.transportista_id == u.id,
                Shipment.date >= start_date,
                Shipment.date <= end_date
            )
            .order_by(Shipment.date.desc(), Shipment.id.desc())
            .all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Detallado"

    headers = [
        "ID", "Fecha Salida", "Fecha Llegada", "Arenera", 
        "Chofer", "DNI", "Patente Tractor", "Patente Batea", "Tipo",
        "Remito Origen", "Tn Origen", 
        "Remito Destino", "Tn Destino (Pagable)", 
        "Merma Descontada (Tn)", 
        "Precio Flete ($)",      
        "Total Neto ($)",        
        "Estado Viaje", "Estado Certificación"
    ]
    ws.append(headers)

    conf = get_config()
    tol_tn = conf.tolerance_kg / 1000.0

    for s in rows:
        is_cert = (s.cert_status == "Certificado")
        
        if s.frozen_flete_neto is not None:
            precio_unit = s.frozen_flete_price or 0
            neto_viaje  = s.frozen_flete_neto or 0
            merma_plata = s.frozen_merma_money or 0
            tn_merma = 0.0
            if merma_plata > 0 and s.frozen_arena_price:
                tn_merma = merma_plata / s.frozen_arena_price
            peso_pagable = (neto_viaje + merma_plata) / (precio_unit if precio_unit else 1)
            remito_final = s.final_remito
        else:
            precio_unit = u.custom_price or 0
            loaded  = s.peso_neto_arenera or 0
            arrived = s.final_peso or s.sbe_peso_neto or 0
            tn_merma = 0.0
            merma_plata = 0.0
            if s.arenera.cert_type == 'salida':
                peso_pagable = loaded
            else:
                diff = loaded - arrived
                if diff > tol_tn:
                    tn_merma = diff - tol_tn
                    merma_plata = tn_merma * (s.arenera.custom_price or 0)
                peso_pagable = arrived - tn_merma
            neto_viaje = (peso_pagable * precio_unit) - merma_plata
            remito_final = s.final_remito if is_cert else (s.sbe_remito or "")

        f_salida = s.date.strftime("%d/%m/%Y")
        f_llegada = s.sbe_fecha_llegada.strftime("%d/%m/%Y") if s.sbe_fecha_llegada else ""
        
        ws.append([
            s.id, f_salida, f_llegada, s.arenera.username if s.arenera else "",
            s.chofer, s.dni, s.tractor, s.trailer, s.tipo,
            s.remito_arenera or "", s.peso_neto_arenera or 0,     
            remito_final or "", peso_pagable or 0,
            tn_merma if tn_merma > 0 else 0, precio_unit, neto_viaje,                      
            s.status, s.cert_status or "Pendiente"
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = (max_length + 2)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = make_response(bio.getvalue())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    fname = f"Reporte_{u.username}_{start_date.strftime('%d%m')}-{end_date.strftime('%d%m')}.xlsx"
    resp.headers["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp

@app.route("/transportista/quotas")
@login_required
@role_required("transportista")
def transportista_quotas():
    u = db.session.get(User, session["user_id"])
    today      = get_arg_today()
    week_dates = [today + timedelta(days=i) for i in range(7)]
    quotas     = (
        Quota.query.filter_by(transportista_id=u.id)
        .filter(Quota.date.in_(week_dates))
        .order_by(Quota.date)
        .all()
    )

    by_date = {}
    for q in quotas:
        if q.arenera:
            by_date.setdefault(q.date, []).append(q)

    totals_by_date = {}
    for d in week_dates:
        lst = by_date.get(d, [])
        totals_by_date[d] = sum(max((q.limit or 0) - (q.used or 0), 0) for q in lst)

    return render_template(
        tpl("transportista_quotas"),
        user=u,
        week_dates=week_dates,
        by_date=by_date,
        totals_by_date=totals_by_date
    )

@app.route("/admin/quotas/arenera/<int:arenera_id>", methods=["GET", "POST"])
@login_required
@role_required("admin", "gestion")
def manage_quotas_arenera(arenera_id):
    # 1. Obtener la Arenera y el rango de fechas
    arenera = User.query.get_or_404(arenera_id)
    transportistas = User.query.filter_by(tipo="transportista").order_by(User.username.asc()).all()

    if request.method == "POST":
        start_str = (request.form.get("start_date") or "").strip()
    else:
        start_str = (request.args.get("start") or "").strip()

    try:
        start_date = date.fromisoformat(start_str) if start_str else get_arg_today()
    except ValueError:
        start_date = get_arg_today()

    week_dates = [start_date + timedelta(days=i) for i in range(7)]

    # 2. Guardar datos (POST)
    if request.method == "POST":
        total_inserts = total_updates = total_deletes = 0
        try:
            for t in transportistas:
                for d in week_dates:
                    # La clave ahora es q_{transportista_id}_{fecha}
                    key = f"q_{t.id}_{d.isoformat()}"
                    raw = (request.form.get(key) or "").strip()

                    # Buscar cupo existente
                    q = db.session.query(Quota).filter_by(
                        transportista_id=t.id,
                        arenera_id=arenera.id, # Arenera fija
                        date=d
                    ).first()

                    if raw == "" or raw == "0":
                        if q:
                            db.session.delete(q)
                            total_deletes += 1
                        continue

                    try:
                        v = int(raw)
                    except ValueError:
                        continue
                    if v < 0: v = 0

                    if q:
                        q.limit = v
                        if (q.used or 0) > v: q.used = v
                        total_updates += 1
                    else:
                        q = Quota(
                            transportista_id=t.id,
                            arenera_id=arenera.id,
                            date=d,
                            limit=v,
                            used=0
                        )
                        db.session.add(q)
                        total_inserts += 1

            db.session.commit()
            flash(f"Cupos actualizados para {arenera.username}: {total_inserts} nuevos, {total_updates} editados.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Error al guardar: {e}", "error")

        return redirect(url_for("manage_quotas_arenera", arenera_id=arenera.id, start=start_date.isoformat()))

    # 3. Preparar datos para mostrar (GET)
    existing = (
        Quota.query
        .filter(
            Quota.arenera_id == arenera.id,
            Quota.date.in_(week_dates)
        ).all()
    )
    # Mapa: (transportista_id, fecha) -> Objeto Quota
    quota_map = {(q.transportista_id, q.date): q for q in existing}

    # Mapa de Uso real (Viajes cargados)
    from sqlalchemy import func
    used_rows = (
        db.session.query(
            Shipment.transportista_id,
            Shipment.date,
            func.count(Shipment.id)
        )
        .filter(
            Shipment.arenera_id == arenera.id,
            Shipment.date >= week_dates[0],
            Shipment.date <= week_dates[-1],
        )
        .group_by(Shipment.transportista_id, Shipment.date)
        .all()
    )
    used_map = {(tid, d): int(c) for (tid, d, c) in used_rows}

    return render_template(
        tpl("admin_quotas_arenera"), # Usaremos una plantilla nueva
        arenera=arenera,
        transportistas=transportistas,
        week_dates=week_dates,
        quota_map=quota_map,
        used_map=used_map,
        start_date=start_date
    )

@app.route("/transportista/arenera/<int:arenera_id>", methods=["GET", "POST"])
@login_required
@role_required("transportista")
def transportista_arenera(arenera_id):
    u = db.session.get(User, session["user_id"])
    date_str = request.args.get("date", get_arg_today().isoformat())
    try: dt = date.fromisoformat(date_str)
    except ValueError: dt = get_arg_today()

    q = Quota.query.filter_by(transportista_id=u.id, arenera_id=arenera_id, date=dt).first_or_404()

    if request.method == "POST":
        if "delete_id" in request.form:
            sid = request.form.get("delete_id")
            s = Shipment.query.get(sid)
            if s and s.transportista_id==u.id and s.status=="En viaje" and s.date==dt:
                db.session.delete(s)
                q.used = max(0, (q.used or 0) - 1)
                db.session.commit()
                flash("Viaje eliminado.", "success")
        else:
            if (q.used or 0) < (q.limit or 0):
                chofer  = request.form.get("nombre_apellido", "").strip()
                dni     = request.form.get("dni", "").strip()
                gender  = request.form.get("gender", "M")
                tipo    = request.form.get("tipo", "")
                tractor = request.form.get("patente_tractor", "").strip().upper().replace(" ", "")
                trailer = request.form.get("patente_batea", "").strip().upper().replace(" ", "")
                
                if chofer and dni and tractor:
                    new_ship = Shipment(
                        transportista_id=u.id,
                        arenera_id=arenera_id,
                        operador_id=u.id,
                        date=dt,
                        chofer=chofer,
                        dni=dni,
                        gender=gender,
                        tipo=tipo,
                        tractor=tractor,
                        trailer=trailer,
                        status="En viaje",
                        peso_neto_arenera=None,
                        cert_status="Pendiente"
                    )
                    db.session.add(new_ship)
                    q.used = (q.used or 0) + 1
                    db.session.commit()
                    flash("Viaje iniciado. Diríjase a la Arenera.", "success")
                else:
                    flash("Faltan datos.", "error")
            else:
                flash("Sin cupo disponible.", "error")
        return redirect(url_for("transportista_arenera", arenera_id=arenera_id, date=dt.isoformat()))

    shipments = Shipment.query.filter_by(transportista_id=u.id, arenera_id=arenera_id, date=dt).order_by(Shipment.id.desc()).all()

    return render_template(
        tpl("transportista_arenera"),
        assignment=q, shipments=shipments, used=(q.used or 0), limit=(q.limit or 0), date=dt
    )

@app.route("/transportista/choferes", methods=["GET", "POST"])
@login_required
@role_required("transportista")
def transportista_choferes():
    u_id = session["user_id"]
    
    if request.method == "POST":
        action = request.form.get("action")
        if action == "save":
            dni = (request.form.get("dni") or "").strip()
            nombre = (request.form.get("nombre") or "").strip()
            tractor = (request.form.get("tractor") or "").strip().upper().replace(" ", "")
            trailer = (request.form.get("trailer") or "").strip().upper().replace(" ", "")
            gender = (request.form.get("gender") or "M").strip()
            tipo = (request.form.get("tipo") or "").strip() 
            chofer_id = request.form.get("chofer_id", type=int)

            if not (dni and nombre):
                flash("El DNI y el Nombre son obligatorios.", "error")
                return redirect(url_for("transportista_choferes"))

            if chofer_id:
                chofer = Chofer.query.filter_by(id=chofer_id, transportista_id=u_id).first()
                if not chofer:
                    flash("Chofer no encontrado.", "error")
                    return redirect(url_for("transportista_choferes"))
            else:
                exists = Chofer.query.filter_by(dni=dni).first()
                if exists:
                    flash(f"Ya existe un chofer con DNI {dni} registrado.", "error")
                    return redirect(url_for("transportista_choferes"))
                chofer = Chofer(transportista_id=u_id, dni=dni)
            
            chofer.nombre = nombre
            chofer.tractor = tractor
            chofer.trailer = trailer
            chofer.gender = gender
            chofer.tipo = tipo

            db.session.add(chofer)
            db.session.commit()
            flash(f"Chofer {nombre} guardado exitosamente.", "success")
        
        elif action == "delete":
            chofer_id = request.form.get("chofer_id", type=int)
            chofer = Chofer.query.filter_by(id=chofer_id, transportista_id=u_id).first_or_404()
            db.session.delete(chofer)
            db.session.commit()
            flash("Chofer eliminado.", "success")

        return redirect(url_for("transportista_choferes"))
    
    choferes_raw = Chofer.query.filter_by(transportista_id=u_id).order_by(Chofer.nombre.asc()).all()
    choferes = []
    for c in choferes_raw:
        choferes.append({
            'id': c.id, 'nombre': c.nombre, 'dni': c.dni,
            'gender': c.gender, 'tractor': c.tractor, 'trailer': c.trailer, 'tipo': c.tipo,
        })
    return render_template(tpl("transportista_choferes"), choferes=choferes)

@app.route("/api/choferes/mine")
@login_required
@role_required("transportista")
def api_choferes_mine():
    u_id = session["user_id"]
    choferes = db.session.query(Chofer).filter_by(transportista_id=u_id).all()
    data = []
    for c in choferes:
        data.append({
            'dni': c.dni, 'nombre': c.nombre, 'gender': c.gender,
            'tractor': c.tractor, 'trailer': c.trailer, 'tipo': c.tipo,
        })
    return data

@app.post("/transportista/shipment/<int:ship_id>/delete")
@login_required
@role_required("transportista")
def delete_own_shipment(ship_id):
    s = Shipment.query.get_or_404(ship_id)
    if s.transportista_id != session.get("user_id"):
        abort(403)
    if s.status != "En viaje":
        flash("Sólo se puede eliminar cuando el viaje está 'En viaje'.", "error")
        return redirect(url_for("transportista_panel"))
    q = Quota.query.filter_by(
        transportista_id=s.transportista_id,
        arenera_id=s.arenera_id,
        date=s.date
    ).first()
    if q and q.used > 0:
        q.used -= 1
    db.session.delete(s)
    db.session.commit()
    flash("Viaje eliminado.", "success")
    return redirect(url_for("transportista_panel"))

@app.route("/arenera")
@login_required
@role_required("arenera")
def arenera_panel():
    u = db.session.get(User, session["user_id"])
    fam = get_family_ids(session["user_id"])
    if not fam: 
        flash("Error de permisos.", "error")
        return redirect(url_for("login"))
    
    sh = request.args.get("search", "").strip().lower()
    trans_filter = request.args.get("trans_filter")
    
    q = Shipment.query.filter(
        Shipment.arenera_id.in_(fam),
        Shipment.status == "En viaje" 
    )

    if sh: 
        q = q.filter(
            func.lower(Shipment.chofer).like(f"%{sh}%") | 
            Shipment.dni.like(f"%{sh}%") | 
            func.lower(Shipment.tractor).like(f"%{sh}%")
        )
    
    if trans_filter and trans_filter != "all":
        try:
            q = q.filter(Shipment.transportista_id == int(trans_filter))
        except ValueError: pass

    ships = q.order_by(Shipment.date.asc(), Shipment.id.asc()).all()
    
    active_trans_ids = db.session.query(Shipment.transportista_id).filter(
        Shipment.arenera_id.in_(fam),
        Shipment.status == "En viaje"
    ).distinct().all()
    
    active_transports = []
    if active_trans_ids:
        t_ids = [r[0] for r in active_trans_ids]
        active_transports = User.query.filter(User.id.in_(t_ids)).order_by(User.username).all()

    today = get_arg_today()
    stats = {
        "en_viaje": Shipment.query.filter(Shipment.arenera_id.in_(fam), Shipment.status=="En viaje", Shipment.date==today).count(),
        "salido": Shipment.query.filter(Shipment.arenera_id.in_(fam), Shipment.status=="Salido a SBE", Shipment.date==today).count(),
        "search": sh,
        "trans_filter": trans_filter
    }
    
    return render_template(
        tpl("arenera_panel"), 
        user=u, 
        shipments=ships, 
        stats=stats, 
        active_transports=active_transports
    )

@app.route("/arenera/history")
@login_required
@role_required("arenera")
def arenera_history():
    u = db.session.get(User, session["user_id"])
    fam = get_family_ids(session["user_id"])

    min_db_date = db.session.query(func.min(Shipment.date)).filter(Shipment.arenera_id.in_(fam)).scalar() or get_arg_today()
    max_db_date = db.session.query(func.max(Shipment.date)).filter(Shipment.arenera_id.in_(fam)).scalar() or get_arg_today()

    start_str = request.args.get("start_date", "")
    end_str   = request.args.get("end_date", "")
    
    start_date = date.fromisoformat(start_str) if start_str else min_db_date
    end_date   = date.fromisoformat(end_str)   if end_str   else max_db_date

    if end_date < start_date:
        start_date, end_date = end_date, start_date

    base_q = Shipment.query.filter(
        Shipment.arenera_id.in_(fam),
        Shipment.date >= start_date,
        Shipment.date <= end_date
    )

    all_in_range = base_q.order_by(Shipment.date.desc(), Shipment.id.desc()).all()
    
    counts = { "total": len(all_in_range), "en_viaje": 0, "salido": 0, "llegado": 0, "certificado": 0 }

    for s in all_in_range:
        if s.cert_status == 'Certificado':
            counts["certificado"] += 1
        elif s.status in ['Llego', 'Llegado a SBE'] or s.sbe_fecha_llegada:
            counts["llegado"] += 1
        elif s.status == 'Salido a SBE' or (s.remito_arenera and not s.sbe_fecha_llegada):
            counts["salido"] += 1
        else:
            counts["en_viaje"] += 1

    final_q = base_q
    status = request.args.get("status", "")
    
    if status:
        if status == "Llegado a SBE":
            final_q = final_q.filter(
                (
                    (Shipment.status == "Llego") | 
                    (Shipment.status == "Llegado a SBE") |
                    (Shipment.sbe_fecha_llegada != None)
                ),
                (Shipment.cert_status != "Certificado")
            )
        elif status == "Salido a SBE":
            final_q = final_q.filter(
                (Shipment.status == "Salido a SBE") |
                (
                    (Shipment.remito_arenera != None) & 
                    (Shipment.remito_arenera != "") & 
                    (Shipment.sbe_fecha_llegada == None) & 
                    (Shipment.status != "En viaje") &
                    (Shipment.cert_status != "Certificado")
                )
            )
        elif status == "En viaje":
             final_q = final_q.filter(Shipment.status == "En viaje")
        elif status == "Certificado":
             final_q = final_q.filter(Shipment.cert_status == "Certificado")
        else:
            final_q = final_q.filter(Shipment.status == status)

    search = request.args.get("search", "").strip().lower()
    if search:
        if search.isdigit():
             final_q = final_q.filter(
                 (Shipment.id == int(search)) |
                 (Shipment.dni.like(f"%{search}%")) |
                 (Shipment.remito_arenera.like(f"%{search}%"))
             )
        else:
            final_q = final_q.filter(
                func.lower(Shipment.chofer).like(f"%{search}%") | 
                func.lower(Shipment.tractor).like(f"%{search}%") |
                func.lower(Shipment.remito_arenera).like(f"%{search}%")
            )

    rows = final_q.order_by(Shipment.date.desc(), Shipment.id.desc()).all()

    return render_template(
        tpl("arenera_history"),
        user=u,
        shipments=rows,
        counts=counts,
        start_date=start_date.isoformat(),
        end_date=end_date.isoformat(),
        status=status,
        search=search
    )

@app.post("/arenera/update/<int:shipment_id>")
@login_required
@role_required('arenera')
def arenera_update(shipment_id):
    s = Shipment.query.get_or_404(shipment_id)
    fam = get_family_ids(session["user_id"])
    
    # Verificación de seguridad: ¿Es su carga?
    if s.arenera_id not in fam: abort(403)
    
    # Verificación de estado: Si ya llegó o se certificó, no se toca.
    if s.cert_status == "Certificado" or s.status in ["Llego", "Llegado a SBE"]:
        flash("El viaje ya ha llegado a destino o está certificado. No se puede editar.", "error")
        return redirect(url_for("arenera_history"))
    
    action = request.form.get('action')

    if action == "confirmar_salida":
        rem_raw = request.form.get('remito_arenera', "").strip()
        peso_raw = request.form.get('peso_neto_arenera', "").strip()

        # --- NORMALIZACIÓN DE REMITO ---
        # 1. Si tiene guion (0001-12345), nos quedamos con lo de la derecha (12345)
        if '-' in rem_raw:
            parts = rem_raw.split('-')
            number_part = parts[-1]
        else:
            number_part = rem_raw
        
        # 2. Quitamos ceros a la izquierda (0012345 -> 12345)
        rem_limpio = number_part.lstrip('0')
        # -------------------------------

        if not rem_limpio or not peso_raw: 
            flash("Faltan datos o el remito es inválido (solo ceros).", "error")
            return redirect(url_for("arenera_panel"))
        
        # Validar duplicados usando el remito YA LIMPIO
        dup = Shipment.query.filter(
            Shipment.arenera_id.in_(fam), 
            Shipment.id != s.id, 
            Shipment.remito_arenera == rem_limpio, # Comparamos contra el limpio
            Shipment.cert_status != 'Certificado'
        ).first()
        
        if dup:
            flash(f"⚠️ El remito {rem_limpio} ya existe en el viaje #{dup.id}.", "error")
            return redirect(url_for("arenera_panel"))

        try: 
            s.peso_neto_arenera = float(peso_raw.replace(",", "."))
        except ValueError: 
            flash("Peso inválido.", "error")
            return redirect(url_for("arenera_panel"))
            
        s.remito_arenera = rem_limpio  # Guardamos la versión limpia
        s.status = "Salido a SBE"
        
        # Limpiamos datos SBE por si hubo un cruce previo incorrecto
        s.sbe_remito = None
        s.sbe_peso_neto = None
        s.cert_status = "Pendiente"
        
        s.operador_id = session["user_id"]
        db.session.commit()
        
        flash(f"✅ Salida confirmada. Remito guardado: {rem_limpio}", "success")
        return redirect(url_for("arenera_panel"))

    elif action == "revertir":
        if s.status == "Salido a SBE":
            s.status = "En viaje"
            s.cert_status = "Pendiente"
            flash("Corrección habilitada: El viaje ha vuelto a Recepción.", "info")
            db.session.commit()
            return redirect(url_for("arenera_panel")) 
        else:
            flash("No se puede revertir este viaje.", "error")
            return redirect(url_for("arenera_history"))

    return redirect(url_for("arenera_panel"))

@app.get("/arenera/export")
@login_required
@role_required("arenera")
def arenera_export():
    u = db.session.get(User, session["user_id"])
    family_ids = get_family_ids(session["user_id"])
    if not family_ids:
        flash("Error identificando cuenta.", "error")
        return redirect(url_for("arenera_panel"))

    # 1. Filtros de Fecha
    today = get_arg_today()
    start_str = request.args.get("start")
    end_str   = request.args.get("end")

    try:
        if start_str:
            start_date = date.fromisoformat(start_str)
        else:
            start_date = today.replace(day=1)
        if end_str:
            end_date = date.fromisoformat(end_str)
        else:
            end_date = today
    except ValueError:
        start_date, end_date = today, today

    # 2. Consulta a la Base de Datos
    rows = (Shipment.query
            .filter(
                Shipment.arenera_id.in_(family_ids),
                Shipment.date >= start_date,
                Shipment.date <= end_date,
            )
            .order_by(Shipment.date.desc(), Shipment.id.desc())
            .all())

    # 3. Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial Logística"

    # --- ENCABEZADOS IDENTICOS A LA IMAGEN ---
    headers = [
        "ID", 
        "Fecha Salida", 
        "Transportista", 
        "Chofer", 
        "Patente", 
        "Remito (Salida)", 
        "Tn Salida", 
        "Fecha Llegada SBE", 
        "Remito SBE", 
        "Tn Llegada SBE", 
        "Estado"
    ]
    ws.append(headers)

    for s in rows:
        # Formato de Fechas
        f_salida = s.date.strftime("%d/%m/%Y")
        f_llegada = s.sbe_fecha_llegada.strftime("%d/%m/%Y") if s.sbe_fecha_llegada else "-"
        
        # Datos SBE (Si está certificado usamos el final, si no, el detectado)
        remito_sbe = s.final_remito if s.final_remito else (s.sbe_remito or "-")
        peso_sbe   = s.final_peso if s.final_peso else (s.sbe_peso_neto or 0)
        if peso_sbe == 0: peso_sbe = "-"

        # Lógica de Estado (Humanamente legible)
        estado_str = "En Viaje"
        if s.cert_status == 'Certificado':
            estado_str = "Certificado"
        elif s.status in ['Llego', 'Llegado a SBE'] or s.sbe_fecha_llegada:
            estado_str = "Llegado SBE"
        elif s.status == 'Salido a SBE' or (s.remito_arenera and not s.sbe_fecha_llegada):
            estado_str = "Salido a SBE"

        # Agregar fila
        ws.append([
            s.id,
            f_salida,
            s.transportista.username if s.transportista else "",
            s.chofer,
            s.tractor, # Patente
            s.remito_arenera or "",
            s.peso_neto_arenera or 0,
            f_llegada,
            remito_sbe,
            peso_sbe,
            estado_str
        ])

    # 4. Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = (max_length + 2)

    # 5. Enviar archivo
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = make_response(bio.getvalue())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    fname = f"Historial_{u.username}_{start_date.strftime('%d%m')}-{end_date.strftime('%d%m')}.xlsx"
    resp.headers["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp

@app.post("/admin/dashboard_data")
@app.get("/admin/export")
@login_required
@role_required("admin")
def admin_export():
    # 1. Filtros opcionales
    start_str = (request.args.get("start") or "").strip()
    end_str   = (request.args.get("end") or "").strip()
    status    = (request.args.get("status") or "").strip()

    q = Shipment.query
    
    # Aplicar filtros de fecha si existen
    if start_str:
        try:
            start_date = date.fromisoformat(start_str)
            q = q.filter(Shipment.date >= start_date)
        except ValueError: pass
    
    if end_str:
        try:
            end_date = date.fromisoformat(end_str)
            q = q.filter(Shipment.date <= end_date)
        except ValueError: pass

    if status:
        q = q.filter(Shipment.status == status)

    # Ordenar cronológicamente
    q = q.order_by(Shipment.date.desc(), Shipment.id.desc())
    
    # Ejecutar consulta
    rows = q.all()

    # 2. Crear Libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial Detallado"

    # 3. Definir Encabezados Enriquecidos
    headers = [
        "ID Viaje",
        "Fecha Salida (Arenera)",
        "Fecha Llegada (SBE)",
        "Arenera (Origen)",
        "Transportista",
        "Chofer",
        "DNI Chofer",
        "Patente Tractor (Decl)",
        "Patente Batea (Decl)",
        "Tipo Camión",
        "Remito Origen",
        "Peso Salida (Tn)",
        "Remito Destino (SBE)",
        "Peso Llegada SBE (Tn)",
        "Estado Actual",
        "Estado Certificación",
        "Fecha Certificación",
        "Observaciones"
    ]
    ws.append(headers)

    # 4. Rellenar Filas
    for s in rows:
        # Formato de Fechas
        f_salida = s.date.strftime("%d/%m/%Y")
        f_llegada = s.sbe_fecha_llegada.strftime("%d/%m/%Y %H:%M") if s.sbe_fecha_llegada else "-"
        f_certif = s.cert_fecha.strftime("%d/%m/%Y") if s.cert_fecha else "-"

        # Nombres de empresas (Manejo de nulos)
        arenera_name = s.arenera.username if s.arenera else "N/A"
        trans_name = s.transportista.username if s.transportista else "N/A"

        # Pesos (Manejo de nulos)
        peso_salida = s.peso_neto_arenera if s.peso_neto_arenera is not None else 0
        peso_llegada = s.sbe_peso_neto if s.sbe_peso_neto is not None else 0
        
        # Remitos
        remito_orig = s.remito_arenera or "-"
        remito_dest = s.sbe_remito or "-" # Usamos sbe_remito o final_remito si prefieres lo certificado

        ws.append([
            s.id,
            f_salida,
            f_llegada,
            arenera_name,
            trans_name,
            s.chofer,
            s.dni,
            s.tractor,
            s.trailer,
            s.tipo,
            remito_orig,
            peso_salida,
            remito_dest,
            peso_llegada,
            s.status,
            s.cert_status,
            f_certif,
            s.observation_reason or ""
        ])

    # 5. Ajuste Automático de Ancho de Columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        # Límite máximo para que no queden columnas gigantes
        if adjusted_width > 50: adjusted_width = 50
        ws.column_dimensions[column].width = adjusted_width

    # 6. Guardar en memoria y enviar
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    
    resp = make_response(bio.getvalue())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    fname = f"Reporte_Admin_{get_arg_now().strftime('%Y%m%d_%H%M')}.xlsx"
    resp.headers["Content-Disposition"] = f'attachment; filename="{fname}"'
    
    return resp

@app.route("/admin/config", methods=["GET", "POST"])
@login_required
@role_required("admin")
def admin_config():
    conf = get_config()
    if request.method == "POST":
        conf.tolerance_kg   = float(request.form.get("tolerance_kg", 0))
        conf.dispatch_price = float(request.form.get("dispatch_price", 0))
        
        for key, val in request.form.items():
            if "_" not in key: continue
            prefix, uid_str = key.split("_", 1)           
            if prefix in ["email", "price", "certtype", "paydays"]: 
                            try:
                                uid = int(uid_str)
                                u = db.session.get(User, uid)
                                if u:
                                    if prefix == "email": u.email = val.strip()
                                    elif prefix == "price": u.custom_price = float(val) if val else 0.0
                                    elif prefix == "certtype": u.cert_type = val
                                    elif prefix == "paydays": u.payment_days = int(val) if val else 30
                            except ValueError:
                                pass
        db.session.commit()
        flash("Configuracion actualizada.", "success")
        return redirect(url_for("admin_config"))
  
    transportistas = User.query.filter_by(tipo="transportista").order_by(User.username).all()
    areneras       = User.query.filter_by(tipo="arenera").order_by(User.username).all()
    
    return render_template(tpl("admin_config"), conf=conf, transportistas=transportistas, areneras=areneras)

@app.route("/admin/resumen")
@login_required
@role_required("admin", "gestion")
def admin_resumen():
    conf = get_config()
    tol_tn = conf.tolerance_kg / 1000.0 
    
    start_str = request.args.get("start")
    end_str   = request.args.get("end")
    today = get_arg_today()
    
    try:
        start_date = date.fromisoformat(start_str) if start_str else today.replace(day=1)
        end_date   = date.fromisoformat(end_str)   if end_str   else today
    except ValueError:
        start_date, end_date = today, today

    all_trans_users = User.query.filter_by(tipo="transportista").order_by(User.username).all()
    all_trans_ids = set()
    trans_names   = {}
    
    for t in all_trans_users:
        all_trans_ids.add(t.id)
        trans_names[t.id] = t.username

    rows = (Shipment.query
            .filter(Shipment.cert_status == "Certificado")
            .filter(Shipment.cert_fecha >= start_date, Shipment.cert_fecha <= end_date)
            .order_by(Shipment.cert_fecha.asc())
            .all())
    
    mat_neto_cert = {}  
    mat_iva_pay   = {}  
    mat_tn_out    = {}  
    mat_tn_in     = {}  
    mat_tn_diff   = {}  
    mat_trucks    = {}  

    for s in rows:
        tid = s.transportista_id
        if tid not in all_trans_ids:
            all_trans_ids.add(tid)
            trans_names[tid] = s.transportista.username
        
        loaded  = s.peso_neto_arenera or 0
        arrived = s.final_peso or 0
        
        if s.frozen_flete_neto is not None:
            neto = s.frozen_flete_neto
            iva  = s.frozen_flete_iva
            diff_tn_visual = 0
            if s.frozen_merma_money > 0 and (s.frozen_arena_price or 0) > 0:
                diff_tn_visual = s.frozen_merma_money / s.frozen_arena_price
            elif s.frozen_merma_money == 0:
                 diff_tn_visual = 0 
        else:
            price_flete = s.transportista.custom_price or 0
            price_arena = s.arenera.custom_price or 0
            diff = loaded - arrived
            merma_money = 0.0
            diff_tn_visual = 0.0
            
            if s.arenera.cert_type != 'salida':
                if diff > tol_tn: 
                    diff_tn_visual = diff - tol_tn
                    merma_money = diff_tn_visual * price_arena
                payable = arrived - diff_tn_visual
            else:
                payable = loaded

            neto = (payable * price_flete) - merma_money
            iva  = max(0, neto * 1.21)

        d_cert = s.cert_fecha or s.date
        days = s.transportista.payment_days or 30
        d_pago = d_cert + timedelta(days=days)
        
        def add_val(mat, d, t, v):
            if d not in mat: mat[d] = {}
            mat[d][t] = mat[d].get(t, 0) + v

        add_val(mat_neto_cert, d_cert, tid, neto)
        add_val(mat_iva_pay,   d_pago, tid, iva)
        add_val(mat_tn_out,    d_cert, tid, loaded)
        add_val(mat_tn_in,     d_cert, tid, arrived)
        add_val(mat_tn_diff,   d_cert, tid, diff_tn_visual)
        add_val(mat_trucks,    d_cert, tid, 1)

    sorted_trans_ids = sorted(list(all_trans_ids), key=lambda x: trans_names[x])
    
    def process_matrix_data(matrix_dict):
        sorted_dates = sorted(matrix_dict.keys())
        result_rows = []
        totals_by_col = {tid: 0.0 for tid in sorted_trans_ids}
        grand_total_sum = 0.0 

        for d in sorted_dates:
            row_data = {"date": d, "cells": [], "row_total": 0.0}
            for tid in sorted_trans_ids:
                val = matrix_dict[d].get(tid, 0.0)
                row_data["cells"].append(val)
                row_data["row_total"] += val
                totals_by_col[tid] += val
            grand_total_sum += row_data["row_total"]
            result_rows.append(row_data)
        return result_rows, totals_by_col, grand_total_sum

    t1, c1, g1 = process_matrix_data(mat_neto_cert)
    t2, c2, g2 = process_matrix_data(mat_iva_pay)
    t3, c3, g3 = process_matrix_data(mat_tn_out)
    t4, c4, g4 = process_matrix_data(mat_tn_in)
    t5, c5, g5 = process_matrix_data(mat_tn_diff)
    t6, c6, g6 = process_matrix_data(mat_trucks)

    return render_template(tpl("admin_resumen"),
        start=start_date, end=end_date,
        trans_ids=sorted_trans_ids,
        trans_names=trans_names,
        tables=[
            {"title": "Monto a Certificar (Neto)", "sub": "Por Fecha Certificación", "rows": t1, "cols": c1, "tot": g1, "fmt": "money", "bg": "bg-green"},
            {"title": "Monto a Pagar (c/IVA)",     "sub": "Por Fecha PAGO", "rows": t2, "cols": c2, "tot": g2, "fmt": "money", "bg": "bg-blue"},
            {"title": "Toneladas Despachadas",    "sub": "Por Fecha Certificación", "rows": t3, "cols": c3, "tot": g3, "fmt": "float", "bg": "bg-yellow"},
            {"title": "Toneladas Recibidas",      "sub": "Por Fecha Certificación", "rows": t4, "cols": c4, "tot": g4, "fmt": "float", "bg": "bg-gray"},
            {"title": "Toneladas Diferencia (Multa)", "sub": "Por Fecha Certificación", "rows": t5, "cols": c5, "tot": g5, "fmt": "float", "bg": "bg-red"},
            {"title": "Cantidad Camiones",          "sub": "Por Fecha Certificación", "rows": t6, "cols": c6, "tot": g6, "fmt": "int",   "bg": "bg-white"},
        ]
    )

@app.route("/admin/control_arena", methods=["GET", "POST"])
@login_required
@role_required("admin", "gestion")
def admin_control_arena():
    aid = request.args.get("arenera_id")
    start_str = request.args.get("start")
    end_str   = request.args.get("end")
    today = get_arg_today()
    
    # 1. Resolver fechas
    if not start_str:
        last_monday = today - timedelta(days=today.weekday() + 7)
        start_date = last_monday
        end_date   = last_monday + timedelta(days=6)
    else:
        try:
            start_date = date.fromisoformat(start_str)
            end_date   = date.fromisoformat(end_str)
        except ValueError:
            start_date = end_date = today

    areneras = User.query.filter_by(tipo="arenera", parent_id=None).order_by(User.username).all()
    shipments = []
    total_tn = 0.0
    total_money = 0.0
    selected_arenera = None

    if aid and aid != "none":
        selected_arenera = User.query.get(int(aid))
        
        # Query base filtrando por la arenera seleccionada
        q = Shipment.query.filter(Shipment.arenera_id == int(aid))
        
        # --- LÓGICA CORREGIDA ---
        if selected_arenera.cert_type == 'salida':
            # CASO A: Cobran por SALIDA (Fecha Remito)
            # NO pedimos que esté 'Certificado'. Solo pedimos que haya salido (tenga peso).
            q = q.filter(
                Shipment.date >= start_date, 
                Shipment.date <= end_date,
                Shipment.peso_neto_arenera != None,
                Shipment.peso_neto_arenera > 0,
                # Filtramos status para asegurar que no sea un viaje recién creado (En viaje) 
                # sino uno que ya confirmaron salida.
                Shipment.status.in_(['Salido a SBE', 'Llego', 'Llegado a SBE', 'Certificado'])
            )
        else:
            # CASO B: Cobran por LLEGADA (Certificación)
            # Aquí SÍ mantenemos la exigencia de que esté Certificado.
            q = q.filter(
                Shipment.cert_status == "Certificado",
                Shipment.date >= start_date,
                Shipment.date <= end_date
            )
        # ------------------------

        shipments = q.order_by(Shipment.date.asc()).all()
        
        # Cálculos
        for s in shipments:
            peso = s.peso_neto_arenera or 0
            
            # Si ya se certificó, tiene precio congelado. Si no (caso Salida pendiente), precio actual.
            precio = s.frozen_arena_price if s.frozen_arena_price is not None else (s.arenera.custom_price or 0)
            
            monto = peso * precio
            total_tn += peso
            total_money += monto
            
            # Guardamos valores temporales para mostrar en el HTML
            s._calc_precio = precio
            s._calc_total_arena = monto

    return render_template(tpl("admin_control_arena"),
                           areneras=areneras,
                           shipments=shipments,
                           selected_arenera=selected_arenera,
                           sel_aid=aid,
                           start=start_date, end=end_date,
                           total_tn=total_tn, total_money=total_money)

@app.route("/admin/control_flete", methods=["GET", "POST"])
@login_required
@role_required("admin", "gestion")
def admin_control_flete():
    tid = request.args.get("transportista_id")
    start_str = request.args.get("start")
    end_str   = request.args.get("end")
    date_mode = request.args.get("mode", "cert") 
    today = get_arg_today()
    
    if not start_str:
        start_date = end_date = today
    else:
        try:
            start_date = date.fromisoformat(start_str)
            end_date   = date.fromisoformat(end_str)
        except ValueError:
            start_date = end_date = today

    transportistas = User.query.filter_by(tipo="transportista").order_by(User.username).all()
    shipments = []
    total_tn = 0.0
    total_neto = 0.0
    total_iva = 0.0
    total_final = 0.0
    selected_trans = None

    if tid and tid != "none":
        selected_trans = User.query.get(int(tid))
        q = Shipment.query.filter(
            Shipment.transportista_id == int(tid),
            Shipment.cert_status == "Certificado"
        )
        if date_mode == 'travel':
            q = q.filter(Shipment.date >= start_date, Shipment.date <= end_date).order_by(Shipment.date.asc())
        else:
            q = q.filter(Shipment.cert_fecha >= start_date, Shipment.cert_fecha <= end_date).order_by(Shipment.cert_fecha.asc())
        
        shipments = q.all()
        conf = get_config()
        tol_tn = conf.tolerance_kg / 1000.0
        
        for s in shipments:
            if s.frozen_flete_neto is not None:
                precio_flete = s.frozen_flete_price or 0
                merma_money  = s.frozen_merma_money or 0
                neto         = s.frozen_flete_neto or 0
                iva_monto = neto * 0.21
            else:
                loaded  = s.peso_neto_arenera or 0
                arrived = s.final_peso or 0
                precio_flete = s.transportista.custom_price or 0
                precio_arena = s.arenera.custom_price or 0
                tn_base = loaded if s.arenera.cert_type == 'salida' else arrived
                merma_money = 0.0
                if s.arenera.cert_type != 'salida':
                    diff = loaded - arrived
                    if diff > tol_tn:
                        merma_money = (diff - tol_tn) * precio_arena

                neto = (tn_base * precio_flete) - merma_money
                iva_monto = max(0, neto * 0.21)

            total_tn   += (s.final_peso or 0)
            total_neto += neto
            total_iva  += iva_monto
            s._calc_price = precio_flete
            s._calc_merma = merma_money
            s._calc_neto  = neto
            s._calc_total = neto + iva_monto

    total_final = total_neto + total_iva

    return render_template(tpl("admin_control_flete"),
                           transportistas=transportistas,
                           shipments=shipments,
                           selected_trans=selected_trans,
                           sel_tid=tid,
                           start=start_date, end=end_date,
                           date_mode=date_mode,
                           total_tn=total_tn, 
                           total_neto=total_neto,
                           total_iva=total_iva,
                           total_final=total_final)

@app.route("/admin/sync_sbe")
@login_required
@role_required("admin")
def sync_sbe():
    from sync_service import run_sbe_sync
    matches, err = run_sbe_sync(db, Shipment)
    if err: 
        flash(err, "error")
    else: 
        flash(f"Sync ok: {matches} cruces realizados.", "success")
    return redirect(request.referrer or url_for("admin_panel"))

@app.route("/admin/sync_sbe_emergency")
@login_required
@role_required("admin")
def sync_sbe_emergency():
    import emergency_sync_patente
    try:
        # Ejecutamos la lógica del script de emergencia
        emergency_sync_patente.run_emergency_sync()
        flash("⚠️ Sync de Emergencia completado. Revisa los cruces 'Observados'.", "warning")
    except Exception as e:
        print(f"Error en sync emergencia: {e}")
        flash("Ocurrió un error en el proceso de emergencia.", "error")
        
    return redirect(request.referrer or url_for("admin_certificacion"))

@app.route("/admin/generate_pdf", methods=["GET"])
@login_required
@role_required("admin", "gestion")
def generate_pdf():
    pdf_bytes, fname, error = _create_pdf_internal(
        request.args.get("target_id"),
        request.args.get("type"),
        request.args.get("start"),
        request.args.get("end"),
        request.args.get("mode", "cert")
    )
    if error:
        flash(error, "error")
        return redirect(request.referrer or url_for('admin_panel'))
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename={fname}'
    return response

def _create_pdf_internal(target_id, target_type, start_str, end_str, date_mode):
    from xhtml2pdf import pisa
    today = get_arg_today()
    try:
        start_date = date.fromisoformat(start_str)
        end_date   = date.fromisoformat(end_str)
    except (ValueError, TypeError):
        start_date = end_date = today

    target_user = User.query.get(int(target_id))
    if not target_user: return None, None, "Usuario no encontrado"

    q = Shipment.query
    
    # ---------------------------------------------------------
    # 1. LÓGICA DE FILTRADO (IGUAL QUE EN EL PANEL DE CONTROL)
    # ---------------------------------------------------------
    
    # ¿Es una Arenera que cobra por SALIDA? (No requiere certificado)
    is_salida_mode = (target_type == 'arenera' and target_user.cert_type == 'salida')

    if is_salida_mode:
        q = q.filter(
            Shipment.arenera_id == target_user.id,
            Shipment.date >= start_date, 
            Shipment.date <= end_date,
            Shipment.peso_neto_arenera != None,
            Shipment.peso_neto_arenera > 0,
            # Aceptamos viajes en curso o finalizados, siempre que hayan salido
            Shipment.status.in_(['Salido a SBE', 'Llego', 'Llegado a SBE', 'Certificado'])
        ).order_by(Shipment.date.asc())
        
    else:
        # Lógica Estándar (Transportistas o Areneras por Llegada) -> Requiere CERTIFICADO
        q = q.filter(Shipment.cert_status == "Certificado")
        
        if target_type == 'transportista':
            q = q.filter(Shipment.transportista_id == target_user.id)
        else:
            q = q.filter(Shipment.arenera_id == target_user.id)

        if date_mode == 'travel':
            q = q.filter(Shipment.date >= start_date, Shipment.date <= end_date).order_by(Shipment.date.asc())
        else:
            q = q.filter(Shipment.cert_fecha >= start_date, Shipment.cert_fecha <= end_date).order_by(Shipment.cert_fecha.asc())

    shipments = q.all()
    if not shipments: return None, None, "No hay datos para el rango seleccionado."

    # ---------------------------------------------------------
    # 2. CÁLCULOS (SOPORTA VIAJES NO CERTIFICADOS)
    # ---------------------------------------------------------
    total_tn = 0.0
    subtotal = 0.0
    descuento_dinero = 0.0
    items = []
    
    # Precio actual (por si el viaje no está congelado)
    ref_price = target_user.custom_price or 0

    for s in shipments:
        # ¿Tiene valores históricos congelados?
        use_frozen = (s.frozen_flete_neto is not None)
        
        # A. PRECIO
        if use_frozen:
            precio_unit = s.frozen_flete_price if target_type == 'transportista' else s.frozen_arena_price
        else:
            precio_unit = ref_price

        # B. PESO Y TOTALES
        merma_linea = 0.0
        
        if target_type == 'transportista':
            # Lógica Flete
            if use_frozen:
                neto_linea = s.frozen_flete_neto
                merma_linea = s.frozen_merma_money or 0
                # Reconstruimos peso aprox para mostrar
                peso_pagable = (neto_linea + merma_linea) / (precio_unit if precio_unit else 1)
            else:
                # Estimación al vuelo (Si es Transportista, generalmente requiere certificado, 
                # pero dejamos esto por seguridad)
                peso_pagable = s.final_peso or s.sbe_peso_neto or 0
                neto_linea = peso_pagable * precio_unit

        else:
            # Lógica Venta Arena
            # En modo Salida, pagamos lo que dice la balanza de origen
            peso_pagable = s.peso_neto_arenera or 0
            neto_linea = peso_pagable * (precio_unit or 0)

        # Acumular
        total_tn += peso_pagable
        subtotal += (neto_linea + merma_linea)
        descuento_dinero += merma_linea
        
        # Fechas para el PDF
        f_llegada_str = "-"
        if s.sbe_fecha_llegada: 
            f_llegada_str = s.sbe_fecha_llegada.strftime("%d/%m")
        
        f_cert_str = "-"
        if s.cert_fecha:
            f_cert_str = s.cert_fecha.strftime("%d/%m")

        items.append({
            "remito": s.final_remito if s.cert_status == 'Certificado' else (s.remito_arenera or s.sbe_remito),
            "f_salida": s.date.strftime("%d/%m"),
            "f_llegada": f_llegada_str,
            "f_certif": f_cert_str,
            "chofer": s.chofer[:18], 
            "patente": s.tractor,
            "peso": peso_pagable,
            "total_linea": neto_linea
        })

    total_neto = subtotal - descuento_dinero
    total_iva_inc = total_neto * 1.21

    # Renderizar HTML
    html = render_template(
        "pdf_template.html",
        target_id=target_user.id,
        tipo_reporte="Flete / Transporte" if target_type == 'transportista' else "Venta de Arena",
        empresa=target_user.username,
        fecha_emision=today.strftime("%d/%m/%Y"),
        periodo_inicio=start_date.strftime("%d/%m/%Y"),
        periodo_fin=end_date.strftime("%d/%m/%Y"),
        items=items,
        total_tn=total_tn,
        precio_unitario=ref_price,
        subtotal=subtotal,
        descuento_dinero=descuento_dinero,
        tn_merma=0, 
        total_neto=total_neto,
        total_iva_inc=total_iva_inc
    )

    pdf_io = io.BytesIO()
    pisa_status = pisa.CreatePDF(io.StringIO(html), dest=pdf_io)
    if pisa_status.err: return None, None, f"Error PDF: {pisa_status.err}"
    pdf_io.seek(0)
    fname = f"Liquidacion_{target_user.username}_{start_date.strftime('%d%m')}.pdf"
    
    return pdf_io.getvalue(), fname, None

@app.route("/admin/preview_send", methods=["GET"])
@login_required
@role_required("admin", "gestion")
def preview_send():
    target_id   = request.args.get("target_id")
    target_type = request.args.get("type")
    start       = request.args.get("start")
    end         = request.args.get("end")
    mode        = request.args.get("mode", "cert")

    target_user = User.query.get_or_404(int(target_id))
    pdf_src = url_for('generate_pdf', target_id=target_id, type=target_type, start=start, end=end, mode=mode)
    
    return render_template(tpl("admin_preview_send"), 
                           user=target_user, 
                           pdf_url=pdf_src,
                           target_id=target_id, target_type=target_type, start=start, end=end, mode=mode)

@app.post("/admin/send_email_action")
@login_required
@role_required("admin")
def send_email_action():
    target_id = request.form.get("target_id")
    target_type = request.form.get("target_type")
    start = request.form.get("start")
    end = request.form.get("end")
    mode = request.form.get("mode")
    
    email_dest = request.form.get("email_dest")
    subject    = request.form.get("subject")
    body       = request.form.get("body")

    if not email_dest:
        flash("El destinatario no tiene email configurado.", "error")
        return redirect(request.referrer)

    pdf_bytes, fname, error = _create_pdf_internal(target_id, target_type, start, end, mode)
    if error or not pdf_bytes:
        flash(f"Error generando el PDF: {error}", "error")
        return redirect(request.referrer)

    try:
        send_email_graph(
            destinatario=email_dest,
            asunto=subject,
            cuerpo=body,
            attachment_bytes=pdf_bytes,
            attachment_name=fname
        )
        flash(f"✅ Liquidación enviada correctamente a {email_dest} (vía Microsoft)", "success")
    except Exception as e:
        print(f"Error enviando mail: {e}")
        flash(f"❌ Error enviando correo: {e}", "error")

    if target_type == 'transportista':
        return redirect(url_for('admin_control_flete', transportista_id=target_id, start=start, end=end, mode=mode))
    else:
        return redirect(url_for('admin_control_arena', arenera_id=target_id, start=start, end=end))

# -----------------------------------------------------------
# MICROSOFT GRAPH MAIL SERVICE
# -----------------------------------------------------------
def get_graph_token_mail():
    tenant_id = os.getenv("GRAPH_TENANT_ID")
    client_id = os.getenv("GRAPH_CLIENT_ID")
    secret    = os.getenv("GRAPH_CLIENT_SECRET")
    
    url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    data = {
        "client_id": client_id,
        "client_secret": secret,
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials"
    }
    
    try:
        resp = requests.post(url, data=data)
        resp.raise_for_status()
        return resp.json().get("access_token")
    except Exception as e:
        print(f"Error obteniendo Token Graph: {e}")
        return None

def send_email_graph(destinatario, asunto, cuerpo, attachment_bytes=None, attachment_name="documento.pdf"):
    token = get_graph_token_mail()
    if not token:
        raise Exception("No se pudo obtener el token de Microsoft Graph.")

    sender = os.getenv("MAIL_SENDER_EMAIL")
    url = f"https://graph.microsoft.com/v1.0/users/{sender}/sendMail"
    
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    message_payload = {
        "subject": asunto,
        "body": {
            "contentType": "HTML", 
            "content": cuerpo.replace("\n", "<br>") 
        },
        "toRecipients": [{"emailAddress": {"address": destinatario}}]
    }

    if attachment_bytes:
        b64_content = base64.b64encode(attachment_bytes).decode("utf-8")
        message_payload["attachments"] = [
            {
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": attachment_name,
                "contentType": "application/pdf",
                "contentBytes": b64_content
            }
        ]

    payload = {"message": message_payload, "saveToSentItems": "true"}
    resp = requests.post(url, headers=headers, json=payload)
    
    if not resp.ok:
        raise Exception(f"Error Graph API: {resp.status_code} - {resp.text}")
    return True

if __name__ == "__main__":
    app.run(host=os.getenv("FLASK_HOST", "0.0.0.0"),
            port=int(os.getenv("FLASK_PORT", "5000")),
            debug=os.getenv("FLASK_DEBUG", "1") == "1")